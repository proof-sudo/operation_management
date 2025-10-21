from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging
from datetime import datetime

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.warning("Le module openpyxl n'est pas installé. Installation requise: pip install openpyxl")
    openpyxl = None

class ProjectImportWizard(models.TransientModel):
    _name = 'project.import.wizard'
    _description = 'Wizard d\'import de projets depuis Excel'

    import_file = fields.Binary(
        string='Fichier Excel',
        required=True
    )
    import_filename = fields.Char(string='Nom du fichier')
    
    update_existing = fields.Boolean(
        string='Mettre à jour les projets existants',
        default=True
    )
    create_missing = fields.Boolean(
        string='Créer les projets manquants',
        default=True
    )
    create_missing_records = fields.Boolean(
        string='Créer les enregistrements manquants (utilisateurs, clients, etc.)',
        default=True,
        help="Si activé, crée automatiquement les utilisateurs, clients et autres enregistrements manquants"
    )
    
    import_log = fields.Text(string='Journal d\'import', readonly=True)
    success_count = fields.Integer(string='Projets créés/mis à jour', readonly=True)
    error_count = fields.Integer(string='Erreurs', readonly=True)
    created_users_count = fields.Integer(string='Utilisateurs créés', readonly=True)
    created_partners_count = fields.Integer(string='Clients créés', readonly=True)
    created_categories_count = fields.Integer(string='Catégories créées', readonly=True)

    def action_import(self):
        """Action principale d'import"""
        if not self.import_file:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))

        if not openpyxl:
            raise UserError(_(
                "Le module 'openpyxl' n'est pas installé. "
                "Veuillez l'installer avec: pip install openpyxl"
            ))

        # Réinitialiser les compteurs
        log_messages = ["=== DÉBUT DE L'IMPORT ==="]
        success_count = 0
        error_count = 0
        created_users_count = 0
        created_partners_count = 0
        created_categories_count = 0

        try:
            # Lecture du fichier avec openpyxl
            file_content = base64.b64decode(self.import_file)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
            sheet = workbook.active
            
            # Lecture des en-têtes
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value else "")
            
            _logger.info(f"En-têtes détectés: {headers}")
            
            # Mapping des colonnes
            col_mapping = self._create_column_mapping(headers)
            
            # Vérifier les colonnes obligatoires
            if col_mapping['name'] is None:
                raise UserError(_("La colonne 'Nom' est obligatoire dans le fichier Excel."))

            # Parcourir les lignes de données
            for row_idx in range(2, sheet.max_row + 1):
                try:
                    row_data = []
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        row_data.append(cell_value)
                    
                    project_vals, created_counts = self._prepare_project_vals(row_data, col_mapping, row_idx)
                    
                    # Mettre à jour les compteurs de création
                    created_users_count += created_counts.get('users', 0)
                    created_partners_count += created_counts.get('partners', 0)
                    created_categories_count += created_counts.get('categories', 0)
                    
                    project_name = project_vals.get('name')
                    if not project_name:
                        log_messages.append(f"⚠️ Ligne {row_idx}: Nom manquant, ligne ignorée")
                        continue

                    # Rechercher le projet existant
                    project = self.env['project.project'].search([
                        ('name', '=', project_name)
                    ], limit=1)
                    
                    if project and self.update_existing:
                        project.write(project_vals)
                        log_messages.append(f"✓ Ligne {row_idx}: '{project_name}' mis à jour")
                        success_count += 1
                    elif self.create_missing:
                        self.env['project.project'].create(project_vals)
                        log_messages.append(f"✓ Ligne {row_idx}: '{project_name}' créé")
                        success_count += 1
                    else:
                        log_messages.append(f"⏭ Ligne {row_idx}: '{project_name}' ignoré")
                        
                except Exception as e:
                    error_count += 1
                    log_messages.append(f"✗ Ligne {row_idx}: Erreur - {str(e)}")
                    _logger.error(f"Erreur ligne {row_idx}: {str(e)}")

            # Résumé final
            log_messages.append(f"\n=== RÉSUMÉ ===")
            log_messages.append(f"Projets traités avec succès: {success_count}")
            log_messages.append(f"Erreurs: {error_count}")
            if self.create_missing_records:
                log_messages.append(f"Utilisateurs créés: {created_users_count}")
                log_messages.append(f"Clients créés: {created_partners_count}")
                log_messages.append(f"Catégories créées: {created_categories_count}")
            log_messages.append("=== FIN DE L'IMPORT ===")

            # Mettre à jour le wizard avec les résultats
            self.write({
                'import_log': '\n'.join(log_messages),
                'success_count': success_count,
                'error_count': error_count,
                'created_users_count': created_users_count,
                'created_partners_count': created_partners_count,
                'created_categories_count': created_categories_count,
            })

            return self._show_result_wizard()

        except Exception as e:
            _logger.error(f"Erreur générale import: {str(e)}")
            raise UserError(_(f"Erreur lors de l'import: {str(e)}"))

    def _create_column_mapping(self, headers):
        """Crée le mapping entre les colonnes Excel et les champs Odoo"""
        mapping = {}
        
        column_mapping = {
            'Nom': 'name',
            'PM': 'user_id',
            'Nature': 'nature',
            'BU': 'bu',
            'Domaine': 'domaine',
            'Revenus': 'revenue_type',
            'Cat Recurrent': 'cat_recurrent',
            'AM': 'am',
            'Presales': 'presales',
            'Date IN': 'date_in',
            'Pays': 'pays',
            'Customer': 'partner_id',
            'Secteur': 'secteur',
            'Description du Projet': 'description',
            'Circuit': 'circuit',
            'SC': 'sc',
            'CAS Build': 'cas_build',
            'CAS Run': 'cas_run',
            'CAS Train': 'cas_train',
            'CAS Sw': 'cas_sw',
            'CAS Hw': 'cas_hw',
            'CAS': 'cas',
            'Statut': 'etat_projet',
            'Update Date': 'write_date',
        }
        
        for excel_col, odoo_field in column_mapping.items():
            if excel_col in headers:
                mapping[odoo_field] = headers.index(excel_col)
            else:
                mapping[odoo_field] = None
                _logger.warning(f"Colonne '{excel_col}' non trouvée dans le fichier")
        
        return mapping

    def _prepare_project_vals(self, row_data, col_mapping, row_num):
        """Prépare les valeurs pour la création/mise à jour du projet"""
        vals = {}
        created_counts = {'users': 0, 'partners': 0, 'categories': 0}
        
        # Champ nom (obligatoire)
        if col_mapping['name'] is not None:
            cell_value = row_data[col_mapping['name']]
            if cell_value is not None:
                vals['name'] = str(cell_value).strip()
        
        # Champs de sélection
        selection_fields = {
            'nature': 'nature',
            'bu': 'bu',
            'domaine': 'domaine', 
            'revenue_type': 'revenue_type',
            'circuit': 'circuit',
            'etat_projet': 'etat_projet'
        }
        
        for odoo_field in selection_fields:
            if col_mapping[odoo_field] is not None:
                cell_value = row_data[col_mapping[odoo_field]]
                if cell_value is not None:
                    cell_value_str = str(cell_value).strip()
                    if cell_value_str and cell_value_str.lower() != 'none' and cell_value_str != '':
                        converted_value = self._convert_selection_value(odoo_field, cell_value_str)
                        if converted_value:
                            vals[odoo_field] = converted_value
        
        # Champs texte
        text_fields = ['cat_recurrent', 'description']
        for field in text_fields:
            if col_mapping.get(field) is not None:
                cell_value = row_data[col_mapping[field]]
                if cell_value is not None:
                    cell_value_str = str(cell_value).strip()
                    if cell_value_str:
                        vals[field] = cell_value_str
        
        # Champs numériques
        numeric_fields = {
            'cas_build': 'cas_build',
            'cas_run': 'cas_run', 
            'cas_train': 'cas_train',
            'cas_sw': 'cas_sw',
            'cas_hw': 'cas_hw',
            'cas': 'cas'
        }
        
        for excel_field, odoo_field in numeric_fields.items():
            if col_mapping[odoo_field] is not None:
                cell_value = row_data[col_mapping[odoo_field]]
                if cell_value is not None:
                    try:
                        if cell_value != '':
                            vals[odoo_field] = float(cell_value)
                    except (ValueError, TypeError):
                        _logger.warning(f"Ligne {row_num}: Valeur numérique invalide pour {excel_field}: {cell_value}")
        
        # Champs dates
        date_fields = {'date_in': 'date_in'}
        for odoo_field in date_fields:
            if col_mapping[odoo_field] is not None:
                cell_value = row_data[col_mapping[odoo_field]]
                if cell_value:
                    try:
                        if isinstance(cell_value, datetime):
                            vals[odoo_field] = cell_value.strftime('%Y-%m-%d')
                        elif isinstance(cell_value, str):
                            date_str = cell_value.strip()
                            if date_str:
                                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                                    try:
                                        date_obj = datetime.strptime(date_str, fmt)
                                        vals[odoo_field] = date_obj.strftime('%Y-%m-%d')
                                        break
                                    except ValueError:
                                        continue
                    except Exception:
                        _logger.warning(f"Ligne {row_num}: Date invalide pour {odoo_field}: {cell_value}")
        
        # Champs relationnels avec création automatique
        relational_fields = {
            'user_id': ('res.users', 'PM'),  # Chef de projet
            'am': ('res.users', 'AM'),       # Account Manager
            'presales': ('res.users', 'Presales'),
            'sc': ('res.users', 'SC'),       # Solutions Consultant
            'partner_id': ('res.partner', 'Customer'),
            'secteur': ('res.partner.category', 'Secteur'),
            'pays': ('res.country', 'Pays'),
        }
        
        for odoo_field, (model, field_label) in relational_fields.items():
            if col_mapping[odoo_field] is not None:
                cell_value = row_data[col_mapping[odoo_field]]
                if cell_value is not None:
                    cell_value_str = str(cell_value).strip()
                    if cell_value_str:
                        record, created = self._get_or_create_record(
                            model, cell_value_str, row_num, field_label
                        )
                        if record:
                            vals[odoo_field] = record.id
                            if created:
                                if model == 'res.users':
                                    created_counts['users'] += 1
                                elif model == 'res.partner':
                                    created_counts['partners'] += 1
                                elif model == 'res.partner.category':
                                    created_counts['categories'] += 1

        return vals, created_counts

    def _get_or_create_record(self, model, search_value, row_num, field_label):
        """Trouve ou crée un enregistrement"""
        try:
            # Recherche d'abord l'enregistrement existant
            record = None
            
            if model == 'res.users':
                record = self.env[model].search([
                    '|', ('name', '=ilike', search_value),
                    ('login', '=ilike', search_value)
                ], limit=1)
                
            elif model == 'res.partner':
                record = self.env[model].search([
                    ('name', '=ilike', search_value)
                ], limit=1)
                
            elif model == 'res.partner.category':
                record = self.env[model].search([
                    ('name', '=ilike', search_value)
                ], limit=1)
                
            elif model == 'res.country':
                record = self.env[model].search([
                    '|', ('name', '=ilike', search_value),
                    ('code', '=ilike', search_value)
                ], limit=1)
            
            # Si trouvé, retourner l'enregistrement
            if record:
                return record, False
            
            # Si non trouvé et création activée, créer l'enregistrement
            if self.create_missing_records:
                _logger.info(f"Création {model}: '{search_value}'")
                
                if model == 'res.users':
                    # Créer un utilisateur avec un login basé sur le nom
                    login = self._generate_login(search_value)
                    record = self.env[model].create({
                        'name': search_value,
                        'login': login,
                        'password': login,  # Mot de passe par défaut
                    })
                    return record, True
                    
                elif model == 'res.partner':
                    record = self.env[model].create({
                        'name': search_value,
                        'company_type': 'company',
                    })
                    return record, True
                    
                elif model == 'res.partner.category':
                    record = self.env[model].create({
                        'name': search_value,
                    })
                    return record, True
                    
                elif model == 'res.country':
                    _logger.warning(f"Ligne {row_num}: Pays non trouvé et création non supportée: '{search_value}'")
            
            return None, False
            
        except Exception as e:
            _logger.error(f"Erreur recherche/création {model} '{search_value}': {str(e)}")
            return None, False

    def _generate_login(self, name):
        """Génère un login à partir d'un nom"""
        # Nettoyer le nom pour créer un login
        login = name.lower().strip()
        login = ''.join(c for c in login if c.isalnum() or c in [' ', '-', '_']).strip()
        login = login.replace(' ', '.').replace('-', '.')
        
        # Vérifier si le login existe déjà
        base_login = login
        counter = 1
        while self.env['res.users'].search([('login', '=', login)], limit=1):
            login = f"{base_login}{counter}"
            counter += 1
            
        return login

    def _convert_selection_value(self, field, value):
        """Convertit les valeurs de sélection depuis Excel vers Odoo"""
        original_value = value
        value = value.strip().lower()
        
        selection_mapping = {
            'nature': {
                'all': 'all',
                'end to end': 'end_to_end',
                'livraison': 'livraison',
                'service pro': 'service_pro',
                'services pro': 'service_pro',
                'incentive': 'all',
                'refacturation': 'all',
            },
            'revenue_type': {
                'one shot': 'oneshot',
                'oneshot': 'oneshot',
                'recurrent': 'recurrent',
            },
            'circuit': {
                'fast': 'fast',
                'fast track': 'fast',
                'normal': 'normal',
            },
            'bu': {
                'ict': 'ict',
                'cloud': 'cloud',
                'cybersecurity': 'cybersecurity',
                'formation': 'formation',
                'security': 'security',
                'iic': 'ict',
                'cybersecurite': 'cybersecurity',
            },
            'domaine': {
                'secured it (sec)': 'secured_it',
                'agile infrastructure & cloud (aic)': 'agile_infrastructure_cloud',
                'modern network integration (mni)': 'modern_network_integration',
                'digital workspace (dws)': 'digital_workspace',
                'expert & managed services - run': 'expert_managed_services_run',
                'expert & managed services - train': 'expert_managed_services_train',
                'expert & managed services - build': 'expert_managed_services_build',
                'expert & managed services - think': 'expert_managed_services_think',
                'datacenter facilities (dcf)': 'datacenter_facilities',
                'business data integration (bdi)': 'business_data_integration',
                'security': 'secured_it',
                'others': 'others',
                'none': 'none'
            },
            'etat_projet': {
                '0-annulé': 'cancelled',
                '1-non démarré': 'non_demarre',
                '3-en cours - bloqué': 'en_cours_bloque',
                '3-en cours - provisionning': 'en_cours_provisionning',
                '3-en cours - production': 'en_cours_production',
                '3-en cours - expedition': 'en_cours_expedition',
                '3-en cours - dedouanement': 'en_cours_dedouanement',
                '3-en cours - atelier technique': 'en_cours_atelier_technique',
                '3-en cours - deploiement': 'en_cours_deploiement',
                '3-en cours - formation': 'en_cours_formation',
                '3-en cours - kick off client': 'en_cours_kickoff_client',
                '3-en cours - standby client': 'en_cours_standby_client',
                '3-en cours - standby technical issue': 'en_cours_standby_technical_issue',
                '3-en cours - attente prérequis': 'en_cours_attente_prerequis',
                '3-en cours - tests et recette': 'en_cours_tests_recette',
                '3-en cours - rli': 'en_cours_rli',
                '4-terminé - attente pv/bl': 'termine_attente_pv_bl',
                '4-terminé - lévée de reserve': 'termine_levee_reserve',
                '4-terminé - pv/bl signé': 'termine_pv_bl_signe',
                '5-facturé - attente df': 'facture_attente_df',
                '5-facturé - attente livraison': 'facture_attente_livraison',
                '5-facturé - prestations en cours': 'facture_prestations_en_cours',
                '6-draft': 'draft',
                '6-dossier indisponible': 'dossier_indisponible',
                '7-cloturé': 'cloture',
                '8-suivi - contrat licence': 'suivi_contrat_licence',
                '8-suivi - contrat mixte': 'suivi_contrat_mixte',
                '8-suivi - contrat de services': 'suivi_contrat_services',
                '9-suspendu': 'suspendu',
                'cloturé': 'cloture',
                'non démarré': 'non_demarre',
                'en cours': 'en_cours_production',
                'terminé': 'termine_pv_bl_signe',
                'facturé': 'facture_attente_df',
                'draft': 'draft',
                'suspendu': 'suspendu',
            }
        }
        
        if field in selection_mapping:
            for excel_val, odoo_val in selection_mapping[field].items():
                if value == excel_val.lower():
                    return odoo_val
        
        # Fallback par défaut
        fallback_values = {
            'nature': 'all',
            'bu': 'ict', 
            'domaine': 'others',
            'etat_projet': 'non_demarre',
            'revenue_type': 'oneshot',
            'circuit': 'normal'
        }
        
        return fallback_values.get(field, value)

    def _show_result_wizard(self):
        """Affiche le wizard avec les résultats"""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Résultat de l\'import',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }